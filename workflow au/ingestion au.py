import string
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook


def read_file(file_path):
    """
    Read from a file and return its contents
    :param file_path: the file path where the file to be read is
    :return: the lines or None if file not found
    """
    try:
        with open(file_path + "ew", "r") as file:
            lines = file.readlines()
    except FileNotFoundError:
        return None
    return lines


def write_to_file(file_path, msg):
    curr_time = datetime.now().strftime('%Y-%m-%d-%H:%M:%S')
    # check if file exists, if it does overwrite it. else create new file
    # if os.path.isfile(file_path):
    #     file_open_mode = "w"
    # else:
    file_open_mode = "a"
    with open(file_path, file_open_mode) as file:
        file.write(curr_time + " - \n" + msg + "\n")
        file.flush()
        os.fsync(file.fileno())


class Table:
    def __init__(self, ingestion_flow, source, env, source_namespace, target_namespace):
        self.metadata = {"tags": []}
        self.ingestion_flow = ingestion_flow
        self.source = source
        self.env = env

        self.source_db, self.source_schema, self.source_table, self.target_db, self.target_schema, self.target_table = \
            source_namespace[0], source_namespace[1], source_namespace[2], target_namespace[0], target_namespace[1], \
            target_namespace[2]
        self.source_count, self.target_count = None, None
        self.comments = []
        self.status = "In-progress"

    def get_your_param_file(self):
        src_db, src_schema, src_table = self.source_db, self.source_schema, self.source_table
        tgt_db, tgt_schema, tgt_table = self.target_db, self.target_schema, self.target_table

        if self.ingestion_flow == "FORWARD":
            # TODO: add connection [Find connection name in Teradata/Oracle]
            param_text = f"\t{self.env}~EDWBQ_CS~{self.source}~{src_db}~{src_schema}~{src_table}" \
                         f"~{tgt_db}~{tgt_schema}~{tgt_table}~TD_BQ_EXTRACT~~~BQ\n"
        else:
            # TODO: add connection [Find connection name in Teradata/Oracle]
            param_text = f"\t{self.env}~BQ2TD_PAR~{self.source}~{src_db}~{src_schema}~{src_table}" \
                   f"~{tgt_db}~{tgt_schema}~{tgt_table}~TD_BQ_EXTRACT~~~BQ\n"

        return param_text

    def get_your_jct(self):
        jct_sql = f"""\n SELECT * FROM BQ_JOB_STREAMS bjs
                        WHERE source_db_name = '{self.source_db}'   
                        AND source_schema = '{self.source_schema}'
                        AND source_table_name = '{self.source_table}'
                        AND target_table_name in '{self.target_table}'
                        UNION ALL
                        SELECT * FROM BQ_JOB_STREAMS bjs
                        WHERE target_db_name = '{self.target_db}'   
                        AND target_schema = '{self.target_schema}'
                        AND target_table_name = '{self.target_table}'"""
        return jct_sql

    def get_your_diy(self):
        diy_sql = f"""SELECT * FROM BQ_DIY_MASTER 
        WHERE 1=1
        AND source_db = '{self.source_db}'
        AND SOURCE_SCHEMA = '{self.source_schema}'
        AND SOURCE_TABLE = '{self.target_table}'
        order by start_time desc """

        return diy_sql

    def get_query_to_convert_jct_to_p(self):
        pass

    def get_query_to_check_if_you_exist_at_source(self):
        sql = ""
        if self.ingestion_flow == "FORWARD":
            sql = f"""
            select *
            from dbc.tables
            where
            database_name = '{self.source_schema}' and TableName = '{self.source_table}' """
        return sql

    def get_query_to_find_count_of_object(self, get_bq_count=False):
        # --> Get the counts of objects at source/target
        src_db, src_schema, src_table = self.source_db, self.source_schema, self.source_table
        tgt_db, tgt_schema, tgt_table = self.target_db, self.target_schema, self.target_table

        sql = ""
        if not get_bq_count:
            # TODO: Connect to teradata/oracle here [to get counts]
            if self.ingestion_flow == "FORWARD":
                db, table = src_schema, src_table
            else:
                db, table = src_schema, src_table
            if self.source == "TERADATA":
                sql = f"\tselect cast(count(*) as BIGINT), '{table}' from {db}.{table} UNION ALL\n"
            elif self.source == "ORACLE":
                sql = f"\tselect count(*), '{table}' from {db}.{table} UNION ALL\n"
        else:
            if self.ingestion_flow == "FORWARD":
                db, schema, table = tgt_db, tgt_schema, tgt_table
            else:
                db, schema, table = src_db, src_schema, src_table
            sql = f"\tselect count(*), '{table}' from {db}.{schema}.{table} UNION ALL\n"
        return sql


class MetadataHandler:
    def __init__(self, clean_data_input, ingestion_flow, source, env):
        self.metadata_columns_for_excel = ['Source Extract DB', 'Source Extract Schema', 'Source Extract Table',
                                                'Target BQ Project ID', 'Target BQ Dataset', 'Target BQ Table', 'Status'
                                                ,'Comments', 'Src count']

        self.main_table_map_to_table_object = None
        self.initialize(clean_data_input, ingestion_flow, source, env)

    def initialize(self, clean_data_input, ingestion_flow, source, env):
        # create metadata structure
        main_table_map_to_table_object = {}
        for val in clean_data_input:

            ingestion_flow, source, env = ingestion_flow, source, env
            source_db, source_schema, source_table, target_db, target_schema, target_table = \
                val[0], val[1], val[2], val[3], val[4], val[5]
            src_namespace, tgt_namespace = [source_schema, source_schema, source_table], \
                [target_db, target_schema, target_table]

            table_object = Table(ingestion_flow, source, env, src_namespace, tgt_namespace)
            main_table_map_to_table_object[target_table] = table_object

        self.main_table_map_to_table_object = main_table_map_to_table_object

    def add_status(self, target_table_name_param, status_param):
        table_object = self.main_table_map_to_table_object[target_table_name_param]
        table_object.status = status_param

    def add_comment(self, target_table_name_param, comment):
        table_object = self.main_table_map_to_table_object[target_table_name_param]
        table_object.comments += comment

    def add_metadata(self, target_table_name_param, key, value_param):
        table_object = self.main_table_map_to_table_object[target_table_name_param]
        table_object.metadata[key].append(value_param)

    def save_yourself_as_excel(self, file_name):
        workbook = load_workbook(filename=file_name)
        sheet = workbook.active

        column_mapping = string.ascii_uppercase
        excel_sheet_columns = self.metadata_columns_for_excel
        assert len(excel_sheet_columns) <= len(column_mapping)

        # add column names to Excel sheet
        row_counter = 1
        for index in range(len(excel_sheet_columns)):
            sheet[column_mapping[index] + str(row_counter)] = excel_sheet_columns[index]
        row_counter += 1

        # add records to Excel sheet
        metadata_dict = self.main_table_map_to_table_object
        for target_table in metadata_dict:
            table_obj = metadata_dict[target_table]

            comment = ""
            for comment_val in table_obj.comments:
                comment += comment_val

            record = [table_obj.source_db, table_obj.source_schema, table_obj.source_table, table_obj.target_db,
                      table_obj.target_schema, table_obj.target_table, table_obj.status, comment,
                      table_obj.source_count, table_obj.target_count]
            for index in range(len(excel_sheet_columns)):
                cell_val = record[index]
                sheet[column_mapping[index] + str(row_counter)] = cell_val
            row_counter += 1

        workbook.save(filename=file_name)
        print("\nLoading of data completed Successfully!")

    def load_from_excel_into_yourself(self):
        pass




class IngestionEngine:
    def __init__(self, raw_data, ingestion_flow, source, env):
        self.meta_handler_object = None
        self.ingestion_flow, self.source, self.env = None, None, None
        self.root_dir = None
        self.log_file = None
        self.clean_inp = None
        self.count_of_tables = None
        self.blacklisted_schemas = []
        self.excel_file_name = None

        self.initialize_yourself(raw_data, ingestion_flow, source, env)
        self.databases_not_present_at_source, self.tables_not_present_at_source = None, None

    def initialize_yourself(self, raw_data, ingestion_flow, source, env):
        self.root_dir = os.getcwd() + "\\"
        curr_time = datetime.now().strftime('%Y-%m-%d')
        self.log_file = self.root_dir + "logs/log-" + str(curr_time) + ".log"
        print("Outputting the status to: ", self.log_file)

        source_types = ["TERADATA"]
        ingestion_flow_types = ["FORWARD", "REVERSE"]
        env_types = ["PRD", "TS1", "TS3", "DV1", "DV3"]

        # check if params given by user are valid
        if (ingestion_flow not in ingestion_flow_types) or (source not in source_types) or (env not in env_types):
            write_to_file(self.log_file, "ERROR: Ingestion flow type or source or environment given is incorrect!")
            raise AssertionError
        self.ingestion_flow, self.source, self.env = ingestion_flow, source, env

        # clean the data and load to metadata object
        raw_data = raw_data.strip()
        raw_file = raw_data.split("\n")
        if self.ingestion_flow == "FORWARD":
            clean_inp = [i.split("\t") for i in raw_file]
        else:
            clean_inp = [raw_file[i].split("\t") for i in range(len(raw_file)) if i != 3]
        self.clean_inp = clean_inp
        self.count_of_tables = len(clean_inp)
        self.meta_handler_object = MetadataHandler(clean_inp, ingestion_flow, source, env)

        msg = f"\nThe ingestion params are: \nFlow: {self.ingestion_flow}\nSource: {self.source}\n" \
              f"Environment: {self.env} \nCount of tables: {self.count_of_tables} \n"
        write_to_file(self.log_file, msg)

        # get input in clean format comprising 6 rows.ie: 3 source & 3 target
        assertion_check = list(filter(lambda x: len(x) != 6, clean_inp))
        code_error = "ERROR: The format of raw input should be in the form src_db\tsrc_schema\tsrc_table" \
                     "\ttgt_db\ttgt_schema\t_tgt_table. THIS DOES NOT MATCH!"
        if not len(assertion_check) == 0:
            write_to_file(self.log_file, code_error)
            raise AssertionError

        msg = f"This is the input provided: \n{raw_file}\n"
        msg += f"This is the cleaned input: \n{clean_inp}"
        msg += "\n"
        write_to_file(self.log_file, msg)

        # create a excel file to store metadata
        self.excel_file_name = self.root_dir + "logs/t_1.xlsx"
        wb = Workbook()
        wb.save(self.excel_file_name)

    def run_pre_checks(self):
        # --> Check if objects to ingested are in blacklist, if yes reject it
        # --> Check if Objects exist at source, if no then reject it

        # 1. Check if src objects are not in blacklist
        found_blacklisted_schemas = set()
        for target_table_name in self.meta_handler_object.main_table_map_to_table_object:
            table_obj = self.meta_handler_object.main_table_map_to_table_object[target_table_name]
            schema = table_obj.source_schema
            if schema in self.blacklisted_schemas:
                found_blacklisted_schemas.add(schema)
                status, comments, tag = "REJECTED", ["The table is rejected as the base schema "
                                                     "is blacklisted"], ["blacklisted"]
                self.meta_handler_object.add_status(target_table_name, status)
                self.meta_handler_object.add_comment(target_table_name, comments)
                self.meta_handler_object.add_metadata(target_table_name, "tags", tag)

        msg = f"Warning: These schema {found_blacklisted_schemas} provided is blacklisted!\n"
        write_to_file(self.log_file, msg)

        # 2. check if src objects are present
        td_unique_databases = set()
        for val in self.clean_inp:
            db = val[1]
            td_unique_databases.add(db)

        msg = "RUN THIS: Validate that these databases are present:\n"
        for td_database in td_unique_databases:
            sql = f"""
            select *
            from dbc.tables
            where
            database_name = '{td_database}' UNION ALL\n"""
            msg += sql
        msg = msg[:-len(' UNION ALL')]

        msg += "\n\nRUN THIS: Validate that these tables are present:\n"
        for target_table in self.meta_handler_object.main_table_map_to_table_object:
            table_obj = self.meta_handler_object.main_table_map_to_table_object[target_table]
            sql = table_obj.get_query_to_check_if_you_exist_at_source()
            msg += sql
        msg = msg[:-len(' UNION ALL')]
        write_to_file(self.log_file, msg)

        # get database and tables not present at source from user running previous query
        databases_not_present_at_source = input("Enter database objects not present at source: ").strip()
        tables_not_present_at_source = input("Enter tables not present at source: ").strip()
        self.databases_not_present_at_source, self.tables_not_present_at_source = \
            databases_not_present_at_source.split("\n"), tables_not_present_at_source.split("\n")

        # add entries to metadata dict if db/table not present at source
        for val in self.meta_handler_object.main_table_map_to_table_object:
            table_obj = self.meta_handler_object.main_table_map_to_table_object[target_table_name]


            db, table = table_obj.source_schema, table_obj.source_table
            if db in databases_not_present_at_source:
                status, comments, tag = "REJECTED", \
                    ["The table is rejected as the base schema not present at source"], \
                    ["schema is not present at source"]
                self.meta_handler_object.add_status(val, status)
                self.meta_handler_object.add_comment(val, comments)
                self.meta_handler_object.add_metadata(val, "tags", tag)

            else:
                if table in tables_not_present_at_source:
                    status, comments, tag = "REJECTED", \
                        ["The table is rejected as the base table not present at source"], \
                        ["Table is not present at source"]
                    self.meta_handler_object.add_status(val, status)
                    self.meta_handler_object.add_comment(val, comments)
                    self.meta_handler_object.add_metadata(val, "tags", tag)

        msg = f"\n\tThese databases are not present at source: {databases_not_present_at_source}\n" \
              f"\tThese tables are not present at source: {tables_not_present_at_source}"
        write_to_file(self.log_file, msg)

    def get_tables_from_metadata_dict(self, get_non_rejected=True):
        table_list = []
        for target_table in self.meta_handler_object.main_table_map_to_table_object:
            status = self.meta_handler_object.main_table_map_to_table_object[target_table].status
            if get_non_rejected:
                if status != 'REJECTED':
                    table_list.append(target_table)
            else:
                if status == 'REJECTED':
                    table_list.append(target_table)
        return table_list

    def get_query_to_find_count_of_object(self, find_count_from="Non BQ platform", table_list=None):
        # --> Get the counts of objects at source/target
        # get only non-rejected tables
        if table_list is None:  # get all not rejected tables
            accepted_table_list = self.get_tables_from_metadata_dict(get_non_rejected=True)
        else:
            accepted_table_list = table_list

        msg = ""
        if find_count_from == "Non BQ platform":
            # TODO: Connect to teradata/oracle here
            for val in self.clean_inp:
                if self.ingestion_flow == "FORWARD":
                    db, table = val[1], val[2]
                else:
                    db, table = val[3], val[4]

                if table in accepted_table_list:
                    if self.source == "TERADATA":
                        sql = f"\tselect cast(count(*) as BIGINT), '{table}' from {db}.{table} UNION ALL\n"
                    elif self.source == "ORACLE":
                        sql = f"\tselect count(*), '{table}' from {db}.{table} UNION ALL\n"
                    msg += sql
        else:
            for val in self.clean_inp:
                if self.ingestion_flow == "FORWARD":
                    db, schema, table = val[3], val[4], val[5]
                else:
                    db, schema, table = val[0], val[1], val[2]

                if table in accepted_table_list:
                    sql = f"\tselect count(*), '{table}' from {db}.{schema}.{table} UNION ALL\n"
                    msg += sql

        if len(msg) >= len(' UNION ALL'):
            msg = msg[:-len(' UNION ALL')]
        return msg

    def update_count_of_object(self):
        # get source counts from user running get_query_to_find_count_of_object query
        # add src count to metadata dict

        table_to_src_count = """""".strip()
        # TODO: Add user input here
        # table_to_src_count = input("Enter the table to source count:")
        table_to_src_count = table_to_src_count.split("\n")
        table_to_src_count = [i.split("\t") for i in table_to_src_count]
        for val in table_to_src_count:
            table_name, src_count = val[0], val[1]
            status = self.meta_handler_object.main_table_map_to_metadata[table_name]['Status']
            if status != 'REJECTED':
                self.meta_handler_object.main_table_map_to_metadata[table_name]["Src count"] = src_count

    def param_load(self, table_list_param):
        non_rejected_tables = self.get_tables_from_metadata_dict(get_non_rejected=True)
        # TODO: add connection

        msg = "Generating param file: \n"
        env, source = self.env, self.source
        todays_date = datetime.today().strftime('%d-%m-%Y')
        param_dir = f"/users/edwadm/param/{todays_date}_rav"
        param_file = f"/users/edwadm/param/{todays_date}_rav/param_1.txt"

        metadata_table_map = self.meta_handler_object.main_table_map_to_table_object
        if self.ingestion_flow == "FORWARD":
            for target_table_name in table_list_param:
                table_obj = metadata_table_map[table_list_param]
                if target_table_name in non_rejected_tables:
                    msg += table_obj.get_your_param_file() + "\n"
            msg += f"\n\tmkdir {param_dir}; cd {param_dir}; vi {param_file}" \
                   f"\n\n\tcat {param_file}" \
                   f"\n\tcd /apps/edwgcpdata/python/bqscripts;" \
                   f"\n\tnohup ./DIY_src2stg_BQ.py -i db -f {param_file} -a rnayakm@cisco.com  -e {env} &"
        else:
            for target_table_name in table_list_param:
                table_obj = metadata_table_map[table_list_param]
                if target_table_name in non_rejected_tables:
                    msg += table_obj.get_your_param_file() + "\n"
            msg += f"\n\tmkdir {param_dir}; cd {param_dir}, vi {param_file} " \
                   f"\n\tcat {param_file}" \
                   f"\n\tcd /apps/edwgcpdata/python/bq2td_scripts/;" \
                   f"\n\tnohup ./BQ2TD.py -i db -f {param_file} -a rnayakm@cisco.com -e PRD -e {env} &"

        return msg

    def check_if_already_ingested(self):
        non_rejected_tables = self.get_tables_from_metadata_dict(get_non_rejected=True)

        msg = f"Changes to JCT: \n NOTE: " \
              f"\n1. Run this in {self.env} environment!"

        if self.ingestion_flow == "FORWARD":
            for val in self.clean_inp:
                src_db, src_schema, src_table = val[0], val[1], val[2]
                tgt_db, tgt_schema, tgt_table = val[3], val[4], val[5]

                if tgt_table in non_rejected_tables:
                    msg += f"\nTable: {tgt_table}\n\n"
                    diy_sql = f"""	SELECT * FROM BQ_DIY_MASTER 
                        WHERE 1=1
                        AND source_schema = '{src_schema}'
                        AND target_table = '{tgt_table}'
                        order by start_time desc;"""

                    jct_sql = f"""\n SELECT * FROM BQ_JOB_STREAMS bjs
                                    WHERE source_db_name = '{src_db}'   
                                    AND source_schema = '{src_schema}'
                                    AND source_table_name = '{src_table}'
                                    AND target_table_name in '{tgt_table}'
                                    and active_ind = 'Y' 
                                    UNION ALL
                                    SELECT * FROM BQ_JOB_STREAMS bjs
                                    WHERE target_db_name = '{tgt_db}'   
                                    AND target_schema = '{tgt_schema}'
                                    AND target_table_name = '{tgt_table}'
                            """
                    msg += diy_sql + f"\n{'='*80}"
        return msg

    def save_yourself_as_excel(self):
        self.meta_handler_object.save_yourself_as_excel(self.excel_file_name)
    def return_db_ingestion_queries(self):
        pass

    def return_jct_ingestion_queries(self):
        pass

# --------------------------------------------- START OF CODE -------------------------------------------------------

RAW_DATA_FILE = """TDTEST	SLSORDDB_TS3	N_POS_TRX_LN_TO_SOL_LINK_TV	gcp-edwstgdata-nprd-84293	TDTEST_SLSORDDB_TS3_STOCKPILE	N_POS_TRX_LN_TO_SOL_LINK_TV
TDTEST	SLSORDDB_TS3	N_XAAS_SO_SBSCR_ITM_SLS_TRX	gcp-edwstgdata-nprd-84293	TDTEST_SLSORDDB_TS3_STOCKPILE	N_XAAS_SO_SBSCR_ITM_SLS_TRX
TDTEST	ETLONLYDB_TS3	EL_RO_BOOKINGS	gcp-edwstgdata-nprd-84293	TDTEST_ETLONLYDB_TS3_STOCKPILE	EL_RO_BOOKINGS
TDTEST	COMREFDB_TS3	MT_PARTNER_HIERARCHY	gcp-edwstgdata-nprd-84293	TDPROD_COMREFDB_STOCKPILE_TS3	MT_PARTNER_HIERARCHY
TDTEST	COMREFDB_TS3	R_FISCAL_MONTH_TO_YEAR	gcp-edwstgdata-nprd-84293	TDPROD_COMREFDB_STOCKPILE_TS3	R_FISCAL_MONTH_TO_YEAR"""


ingestion_engine_obj = IngestionEngine(raw_data=RAW_DATA_FILE, ingestion_flow="FORWARD", source="TERADATA", env="TS3")


print("\n--------- STAGE 1: Perform pre checks ---------------")
# --> Check if objects to ingested are in blacklist
# --> Check if Objects exist at source
ingestion_engine_obj.run_pre_checks()

# --> Get the counts of objects at source

msg = "\n--------- STAGE 2: Get count of objects at source ---------------"
msg += "\nRUN THIS: Validate that these tables have counts > 0\n"
source_count_query = ingestion_engine_obj.get_query_to_find_count_of_object()
write_to_file(ingestion_engine_obj.log_file, msg + source_count_query)
# ingestion_engine_obj.update_count_of_object()


msg = "\n--------- STAGE 3: Trigger Ingestion ---------------\n"
check_if_already_ingested = ingestion_engine_obj.check_if_already_ingested()
msg += check_if_already_ingested
write_to_file(ingestion_engine_obj.log_file, msg)

# Enter the objects found in JCT, check if tgt counts match
table_in_jct_count_check = input("Enter the objects found in JCT:")
table_in_jct_count_check = table_in_jct_count_check.strip().split("\t")
table_in_jct_count_check = [i.strip() for i in table_in_jct_count_check]
BQ_count_query = ingestion_engine_obj.get_query_to_find_count_of_object(find_count_from="BQ",
                                                                        table_list=table_in_jct_count_check)
msg = BQ_count_query
write_to_file(ingestion_engine_obj.log_file, msg)


# Enter the history load tables
history_load_tables = input("Enter the history load tables:")
history_load_tables = history_load_tables.split("\t")
history_load_tables = [i.strip() for i in history_load_tables]
history_load_query = ingestion_engine_obj.param_load(history_load_tables)
msg += history_load_query
write_to_file(ingestion_engine_obj.log_file, msg)

ingestion_engine_obj.save_yourself_as_excel()








# # Step-3: Check if they have already been ingested
# # check in DIY and JCT
# # --> if JCT has no entries then history ingestion
# # --> if JCT has 1 entry then either past run failed or src count = 0
# # --> if JCT has 2 entries either past run completed or failed
# run_step_3 = True
# if run_step_3:
#     msg += "\nRUN THIS: Check if they have been ingested previously > 0\n"
#     sql = f"""   SELECT TARGET_TABLE_NAME, count(*) AS COUNT_OF_JCT_ENTRIES
#         ,CASE
#         WHEN count(*) > 2 THEN 'ANOMALY'
#         WHEN count(*) = 2 THEN 'Ingestion attempted'
#         WHEN count(*) = 1 THEN 'Ingestion attempted'
#         WHEN count(*) = 0 THEN 'History load required, not ingested previously'
#         ELSE 'The quantity is under 30'
#         END AS INGESTION_MODE
#        from
#        ("""
#     for val in clean_inp:
#         src_db, src_schema, src_table, tgt_db, tgt_schema, tgt_table = val[0], val[1], val[2], val[3], val[4], \
#             val[5]
#         sql += f"""
#              SELECT * FROM EJC.BQ_JOB_STREAMS
#              WHERE 1=1
#              AND SOURCE_DB_NAME = '{src_db}'
#              AND SOURCE_SCHEMA = '{src_schema}'
#              AND SOURCE_TABLE_NAME = '{src_table}'
#              UNION ALL
#              SELECT * FROM EJC.BQ_JOB_STREAMS
#              WHERE 1=1
#              AND TARGET_DB_NAME = '{tgt_db}'
#              AND TARGET_SCHEMA = '{tgt_schema}'
#              AND TARGET_TABLE_NAME = '{tgt_table}'
#             UNION ALL"""
#     sql = sql[:-len(' UNION ALL')]
#     sql += """)
#            GROUP BY TARGET_TABLE_NAME"""
#     print(sql)
#
#     for val in clean_inp:
#         src_db, src_schema, src_table, tgt_db, tgt_schema, tgt_table = val[0], val[1], val[2], val[3], val[4], \
#             val[5]
#
#         sql_2 = f"""                     SELECT * FROM EJC.BQ_DIY_MASTER
#                  WHERE 1=1
#                  AND INGESTION_TYPE IN ('JCT', 'DB')
#                  AND SOURCE_DB_NAME = '{src_db}'
#                  AND SOURCE_SCHEMA = '{src_schema}'
#                  AND SOURCE_TABLE  = '{src_table}'
#                  AND ENV_TYPE = '{ENV}'
#                  UNION ALL
#                  SELECT * FROM EJC.BQ_DIY_MASTER
#                  WHERE 1=1
#                  AND NOT INGESTION_TYPE IN ('JCT', 'DB')
#                  AND TARGET_SCHEMA = '{tgt_schema}'
#                  AND TARGET_TABLE = '{tgt_table}'
#                  AND ENV_TYPE = '{ENV}'
#                  UNION ALL
#                 """
#         sql_2 = sql_2[:-len(' UNION ALL')]
#         # print(sql_2)
#
#     # separate into already ingested and not ingested:
#     # --> if already ingested check if run is success
#     #       >> if success: update JCT. if failed: Get DIY error
#     # --> if not already ingested: run in param mode
#     ALREADY_INGESTED_AND_FAILED_TABLES = input("Enter ingested & failed tables: ").strip()
#     ALREADY_INGESTED_AND_SUCCEEDED_TABLES = input("Enter ingested & completed tables: ").strip()
#     FIRST_TIME_INGESTION_TABLES = input("Enter not ingested tables: ").strip()
#     ALREADY_INGESTED_AND_FAILED_TABLES, ALREADY_INGESTED_AND_SUCCEEDED_TABLES, FIRST_TIME_INGESTION_TABLES = \
#         ALREADY_INGESTED_AND_FAILED_TABLES.split("\n"), ALREADY_INGESTED_AND_SUCCEEDED_TABLES.split("\n"), \
#         FIRST_TIME_INGESTION_TABLES.split("\n")
#
#     all_tables_list = [val[2] for val in clean_inp]
#
#     boolean_proposition = """len(all_tables_list) == len(ALREADY_INGESTED_AND_FAILED_TABLES) \
#                + len(ALREADY_INGESTED_AND_SUCCEEDED_TABLES) \
#                + len(FIRST_TIME_INGESTION_TABLES)"""
#     CODE_ERROR = "ERROR: The count of provided tables: Not Ingested + Ingested and completed + Ingested and " \
#                  "failed is not consistent"
#     assert_proposition(boolean_proposition=boolean_proposition, file_assertion_error_msg=CODE_ERROR)
#
# # Step-4: Trigger runs
# # --> If Not ingested: Run in param mode
# # --> If Already Ingested: Changed JCT's to P and rerun
# # --> If Ingestion attempted, and it failed --> Check the error
# run_step_4 = True
# if run_step_4:
#     pass
#
# # Step-5: Monitor runs
# # --> Queuing system for large loads [Run plan]
# # --> On completion update the dataframe
# # --> On failure alert me
# run_step_5 = True
# if run_step_5:
#     pass
#
# if ENV == "PRD":
#     pass
# elif ENV == "TS1" or ENV == "TS3":
#     pass
# elif ENV == "DV1" or ENV == "DV3":
#     pass


